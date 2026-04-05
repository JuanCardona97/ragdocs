"""
Document Processor — Extracts text from files and splits into chunks.

Supports PDF, DOCX, TXT, Markdown, CSV, and Excel files.
Uses LangChain's RecursiveCharacterTextSplitter for intelligent chunking.
"""

import csv
import io

import pymupdf
from langchain.text_splitter import RecursiveCharacterTextSplitter


class DocumentProcessor:
    """Handles text extraction from various file formats and chunking."""

    def extract_text(self, content: bytes, file_type: str) -> str:
        """Extract raw text from a document based on its file type."""
        extractors = {
            ".pdf": self._extract_pdf,
            ".docx": self._extract_docx,
            ".txt": self._extract_plain,
            ".md": self._extract_plain,
            ".csv": self._extract_csv,
            ".xlsx": self._extract_xlsx,
        }

        extractor = extractors.get(file_type)
        if not extractor:
            raise ValueError(f"Unsupported file type: {file_type}")

        text = extractor(content)

        if not text.strip():
            raise ValueError("No text could be extracted from the document")

        return text

    def split_into_chunks(
        self,
        text: str,
        chunk_size: int = 1000,
        chunk_overlap: int = 200,
    ) -> list[str]:
        """Split text into overlapping chunks for embedding."""
        splitter = RecursiveCharacterTextSplitter(
            chunk_size=chunk_size,
            chunk_overlap=chunk_overlap,
            length_function=len,
            separators=["\n\n", "\n", ". ", " ", ""],
        )
        return splitter.split_text(text)

    def _extract_pdf(self, content: bytes) -> str:
        """Extract text from a PDF file."""
        doc = pymupdf.open(stream=content, filetype="pdf")
        text_parts = [page.get_text() for page in doc]
        doc.close()
        return "\n\n".join(text_parts)

    def _extract_docx(self, content: bytes) -> str:
        """Extract text from a DOCX file."""
        from docx import Document

        doc = Document(io.BytesIO(content))
        return "\n\n".join(p.text for p in doc.paragraphs if p.text.strip())

    def _extract_plain(self, content: bytes) -> str:
        """Extract text from plain text or markdown files."""
        return content.decode("utf-8", errors="ignore")

    def _extract_csv(self, content: bytes) -> str:
        """Extract text from a CSV file, converting rows to readable text."""
        text_content = content.decode("utf-8", errors="ignore")
        reader = csv.reader(io.StringIO(text_content))
        rows = list(reader)

        if not rows:
            return ""

        headers = rows[0]
        lines = []
        for row in rows[1:]:
            pairs = [f"{h}: {v}" for h, v in zip(headers, row) if v.strip()]
            if pairs:
                lines.append(" | ".join(pairs))

        return "\n".join(lines)

    def _extract_xlsx(self, content: bytes) -> str:
        """Extract text from an Excel file, converting each sheet to text."""
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        all_text = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue

            all_text.append(f"--- Sheet: {sheet_name} ---")
            headers = [str(c) if c is not None else "" for c in rows[0]]

            for row in rows[1:]:
                pairs = []
                for h, v in zip(headers, row):
                    if v is not None and str(v).strip():
                        pairs.append(f"{h}: {v}")
                if pairs:
                    all_text.append(" | ".join(pairs))

        wb.close()
        return "\n".join(all_text)
